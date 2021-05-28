# -*- coding: utf-8 -*-
"""
Created on Tue May 30 14:53:46 2020

@author: 郝蛤蛤

.　　   　  ▃▆█▇▄▖
　　   　▟◤▖　　　◥█▎
　　　◢◤　 ▐　　　　▐▉
　 ▗◤　　　▂　▗▖　　▕█▎
  ◤　▗▅▖◥▄　▀◣　　█▊
▐　▕▎◥▖◣◤　　　　◢██
█◣　◥▅█▀　　　　▐██◤
▐█▙▂　　　   ◢██◤
　◥██◣　　◢▄◤
　　　▀██▅▇▀
哼哼，啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊
"""
import sys
import openpyxl as pyxl
import matplotlib.pyplot as plt
#全局字体设置为Times New Roman
plt.rc('font', family='Times New Roman')

filename = sys.argv[1]
workbook = pyxl.load_workbook(filename)
shenames = workbook.get_sheet_names()
list_line4 = [[], [], []]

def print_XY0(name, label, columns):

    # data_x_list = listx #例如48601-RMSD-potein-potein-xvgx
    data_x_list = range(0, columns - 2)
    fig = plt.figure(figsize=(10, 5))

    DM1 = fig.add_subplot(1, 1, 1)

    DM1.plot(data_x_list, data1_y_list, 'm-', linewidth=0.5, label=label)  # y1轴
    DM1.legend(frameon=False)
    # DM1.legend(frameon=False)  # 图例，去掉框框
    plt.xlim(0, columns - 1)  # 坐标轴范围
    DM1.set_xlabel("Residue") #例如Time(ns)
    DM1.set_ylabel("Energy_" + name + "(KJ/mol)") #例如RMSD(ns)

    plt.show()

def print_XY1(listy, rows_last_one):
    data_x_list = range(1, rows_last_one) #例如48601-RMSD-potein-potein-xvgx
    data1_y_list = listy #例如48601-RMSD-potein-potein-xvgy1

    fig = plt.figure(figsize=(10, 5))

    DM1 = fig.add_subplot(1, 1, 1)

    DM1.plot(data_x_list, data1_y_list, 'm-', label="Binding_Energy", linewidth=0.5)  # y1轴

    DM1.legend(frameon=False)  # 图例，去掉框框
    plt.xlim(0, rows_last_one)  # 坐标轴范围
    DM1.set_xlabel("Residue") #例如Time(ns)
    DM1.set_ylabel("Energy(KJ/mol)") #例如RMSD(ns)

    plt.show()
#遍历前三个，读取数据
for num in range(len(shenames) - 3):
    sheet = workbook.get_sheet_by_name(shenames[num])
    rows = sheet.max_row
    columns = sheet.max_column
    print("总行数为：" + str(rows), "总列数为：" + str(columns))
    #读取第四行的数据
    for i in range(2, columns):
        line4 = sheet.cell(row=rows, column=i).value
        list_line4[num].append(line4)

    data1_y_list = list_line4[num]
    print("正在绘制第" + str(num + 1) +"张图，名字叫" + str(shenames[num]))
    name = str(shenames[num])
    name = name.replace(name[0:8], "")
    name = name.replace(name[-4:], "")
    label = str(shenames[num])
    label = label.replace(label[-4:], "")
    print_XY0(name, label, columns)
#读取最后一个sheet
sheet_last_one = workbook.get_sheet_by_name(shenames[-1])
rows_last_one = sheet_last_one.max_row
columns_last_one = sheet_last_one.max_column
print(rows_last_one, columns_last_one)
listy = []
#遍历所有
for j in range(1, rows_last_one):
    row1 = sheet_last_one.cell(row=j, column=2).value
    listy.append(row1)
print_XY1(listy, rows_last_one)

